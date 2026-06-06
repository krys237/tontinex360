"""
Service d'import massif de membres depuis un fichier Excel.

Deux modes :
- DIRECT : crée immédiatement les Memberships actifs (et User si nécessaire)
- INVITE : envoie une Invitation par ligne (l'utilisateur doit accepter)

Format Excel attendu (en-têtes flexibles, reconnaissance souple) :
    telephone | first_name | last_name | email | member_number
    (telephone obligatoire, le reste optionnel)

Synonymes acceptés pour les en-têtes :
    telephone : ['telephone', 'téléphone', 'phone', 'tel', 'mobile', 'numero']
    first_name: ['first_name', 'prenom', 'prénom', 'firstname']
    last_name : ['last_name', 'nom', 'lastname', 'surname']
    email     : ['email', 'mail', 'e-mail']
    member_number: ['member_number', 'matricule', 'numero_membre', 'n°', 'code']
"""
import re
from datetime import timedelta
from io import BytesIO
from typing import Iterable

from django.db import transaction as db_transaction
from django.utils import timezone
from openpyxl import load_workbook


# ─── Mapping des colonnes ───────────────────────────────────────────


HEADER_SYNONYMS = {
    'telephone': ['telephone', 'téléphone', 'phone', 'tel', 'mobile', 'numero', 'numéro'],
    'first_name': ['first_name', 'firstname', 'prenom', 'prénom', 'first'],
    'last_name': ['last_name', 'lastname', 'nom', 'surname', 'last', 'family_name'],
    'email': ['email', 'mail', 'e-mail', 'courriel'],
    'member_number': ['member_number', 'matricule', 'numero_membre', 'numéro_membre',
                      'n°', 'no', 'code', 'membership_number'],
}


def _normalize_header(h: str) -> str:
    if h is None:
        return ''
    s = str(h).strip().lower()
    s = s.replace('-', '_').replace(' ', '_')
    s = re.sub(r'[^a-z0-9_éèàâêïôûç]', '', s)
    return s


def _match_column(normalized_header: str) -> str | None:
    for canonical, synonyms in HEADER_SYNONYMS.items():
        if normalized_header in [_normalize_header(s) for s in synonyms]:
            return canonical
    return None


# ─── Normalisation des valeurs ──────────────────────────────────────


PHONE_RE = re.compile(r'[^\d+]')


def _normalize_phone(raw) -> str:
    if raw is None:
        return ''
    s = str(raw).strip()
    # Garde + et chiffres
    s = PHONE_RE.sub('', s)
    if not s:
        return ''
    # Format canonique : si commence par 0, on garde tel quel (local). Si pas de +, on ajoute pas
    # (l'association peut être au Cameroun/Sénégal/etc., on ne présume pas).
    return s


def _normalize_email(raw) -> str:
    if raw is None:
        return ''
    s = str(raw).strip().lower()
    return s if '@' in s else ''


def _normalize_name(raw) -> str:
    if raw is None:
        return ''
    return str(raw).strip()[:100]


# ─── Parsing ────────────────────────────────────────────────────────


def parse_excel(file_bytes: bytes) -> tuple[list[dict], list[str]]:
    """
    Parse le contenu binaire d'un .xlsx en lignes normalisées.
    Retourne (rows, errors). Une row = {row_number, raw_data, parsed_*}.
    """
    errors: list[str] = []
    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        return [], [f"Fichier illisible : {e}"]

    sheet = wb.active
    if sheet is None:
        return [], ["Aucune feuille trouvée dans le classeur."]

    rows_iter = sheet.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], ["Le fichier est vide."]

    # Mapper position -> nom canonique
    col_map: dict[int, str] = {}
    for idx, h in enumerate(header_row):
        canonical = _match_column(_normalize_header(h or ''))
        if canonical:
            col_map[idx] = canonical

    if 'telephone' not in col_map.values():
        return [], [
            "Aucune colonne 'telephone' détectée. En-têtes acceptés : "
            + ", ".join(HEADER_SYNONYMS['telephone'])
        ]

    parsed_rows: list[dict] = []
    for row_idx, row in enumerate(rows_iter, start=2):  # start=2 (en-tête = ligne 1)
        if not any(row):  # ligne vide
            continue
        raw = {}
        parsed = {
            'telephone': '', 'first_name': '', 'last_name': '',
            'email': '', 'member_number': '',
        }
        for col_idx, value in enumerate(row):
            if value is None:
                continue
            canonical = col_map.get(col_idx)
            if canonical:
                raw[canonical] = str(value).strip()
                if canonical == 'telephone':
                    parsed['telephone'] = _normalize_phone(value)
                elif canonical == 'email':
                    parsed['email'] = _normalize_email(value)
                elif canonical in ('first_name', 'last_name'):
                    parsed[canonical] = _normalize_name(value)
                else:
                    parsed[canonical] = str(value).strip()[:100]

        parsed_rows.append({
            'row_number': row_idx,
            'raw_data': raw,
            **{f'parsed_{k}': v for k, v in parsed.items()},
        })

    return parsed_rows, errors


# ─── Validation ─────────────────────────────────────────────────────


def validate_rows(association, parsed_rows: list[dict]) -> list[dict]:
    """
    Annote chaque ligne d'une `validation` :
    - 'ok'       : prête à être importée
    - 'duplicate': téléphone déjà présent comme membre actif
    - 'invalid'  : téléphone manquant ou invalide
    - 'doublon_fichier' : téléphone répété dans le même fichier
    """
    from apps.members.models import Membership

    seen_in_file: dict[str, int] = {}
    existing_phones = set(
        Membership.all_objects.filter(
            association=association, is_active=True,
        ).values_list('user__telephone', flat=True),
    )

    for row in parsed_rows:
        phone = row['parsed_telephone']
        if not phone or len(phone) < 6:
            row['validation'] = 'invalid'
            row['validation_message'] = "Téléphone manquant ou invalide."
            continue
        if phone in seen_in_file:
            row['validation'] = 'doublon_fichier'
            row['validation_message'] = (
                f"Doublon avec la ligne {seen_in_file[phone]} du fichier."
            )
            continue
        seen_in_file[phone] = row['row_number']
        if phone in existing_phones:
            row['validation'] = 'duplicate'
            row['validation_message'] = "Membre actif déjà existant pour ce téléphone."
            continue
        row['validation'] = 'ok'
        row['validation_message'] = ''
    return parsed_rows


# ─── Génération de template Excel ──────────────────────────────────


def build_template_xlsx() -> bytes:
    """Renvoie un .xlsx prérempli avec les bonnes colonnes + une ligne exemple."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Membres"
    ws.append(['telephone', 'first_name', 'last_name', 'email', 'member_number'])
    ws.append(['237699999999', 'Jean', 'Dupont', 'jean@example.com', 'M-001'])
    ws.append(['+225 07 08 09 10 11', 'Marie', 'NDIAYE', '', ''])
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─── Traitement (DIRECT + INVITE) ──────────────────────────────────


@db_transaction.atomic
def process_batch(batch, parsed_rows: list[dict]) -> dict:
    """
    Crée les MemberImportRow + applique le mode (DIRECT ou INVITE).
    Idempotent : on saute les rows déjà validées dans des batches existants
    via les checks de doublon.
    """
    from apps.members.models import MemberImportBatch, MemberImportRow

    batch.status = MemberImportBatch.Status.PROCESSING
    batch.total_rows = len(parsed_rows)
    batch.save(update_fields=['status', 'total_rows', 'updated_at'])

    success = 0
    error = 0
    skipped = 0

    for row in parsed_rows:
        import_row = MemberImportRow.all_objects.create(
            association=batch.association,
            batch=batch,
            row_number=row['row_number'],
            raw_data=row.get('raw_data', {}),
            parsed_telephone=row['parsed_telephone'],
            parsed_first_name=row['parsed_first_name'],
            parsed_last_name=row['parsed_last_name'],
            parsed_email=row['parsed_email'],
            parsed_member_number=row['parsed_member_number'],
        )

        validation = row.get('validation', 'ok')
        if validation == 'invalid':
            import_row.status = MemberImportRow.Status.ERROR
            import_row.error_message = row.get('validation_message', 'Ligne invalide.')
            import_row.save(update_fields=['status', 'error_message', 'updated_at'])
            error += 1
            continue
        if validation in ('duplicate', 'doublon_fichier'):
            import_row.status = MemberImportRow.Status.SKIPPED
            import_row.error_message = row.get('validation_message', '')
            import_row.save(update_fields=['status', 'error_message', 'updated_at'])
            skipped += 1
            continue

        try:
            if batch.mode == MemberImportBatch.Mode.DIRECT:
                _process_direct(batch, import_row)
            else:
                _process_invite(batch, import_row)
            import_row.status = MemberImportRow.Status.SUCCESS
            import_row.save(update_fields=[
                'status', 'resulting_membership', 'resulting_invitation_id',
                'updated_at',
            ])
            success += 1
        except Exception as e:  # noqa: BLE001
            import_row.status = MemberImportRow.Status.ERROR
            import_row.error_message = str(e)[:500]
            import_row.save(update_fields=['status', 'error_message', 'updated_at'])
            error += 1

    batch.status = MemberImportBatch.Status.COMPLETED
    batch.success_count = success
    batch.error_count = error
    batch.skipped_count = skipped
    batch.processed_at = timezone.now()
    batch.save(update_fields=[
        'status', 'success_count', 'error_count', 'skipped_count',
        'processed_at', 'updated_at',
    ])

    return {
        'total': batch.total_rows,
        'success': success,
        'error': error,
        'skipped': skipped,
    }


def _process_direct(batch, import_row):
    """Crée le User (si nécessaire) + le Membership actif immédiatement."""
    from apps.core.models import User
    from apps.members.models import Membership, MemberRole, Role

    user, _ = User.objects.get_or_create(
        telephone=import_row.parsed_telephone,
        defaults={
            'first_name': import_row.parsed_first_name or None,
            'last_name': import_row.parsed_last_name or None,
            'email': import_row.parsed_email or None,
            'is_active': True,
        },
    )

    # Si le user existe sans prénom/nom et qu'on en a, on complète
    updated = False
    if not user.first_name and import_row.parsed_first_name:
        user.first_name = import_row.parsed_first_name
        updated = True
    if not user.last_name and import_row.parsed_last_name:
        user.last_name = import_row.parsed_last_name
        updated = True
    if not user.email and import_row.parsed_email:
        user.email = import_row.parsed_email
        updated = True
    if updated:
        user.save()

    # Vérifier qu'il n'existe pas déjà un membership (même inactif)
    membership = Membership.all_objects.filter(
        association=batch.association, user=user,
    ).first()

    if membership:
        # Réactiver si inactif
        if not membership.is_active:
            membership.is_active = True
            membership.status = Membership.Status.ACTIVE
            membership.save(update_fields=['is_active', 'status', 'updated_at'])
    else:
        membership = Membership.all_objects.create(
            association=batch.association,
            user=user,
            status=Membership.Status.ACTIVE,
            is_active=True,
            member_number=import_row.parsed_member_number or '',
        )

    # Rôle 'membre' par défaut
    default_role = Role.all_objects.filter(
        association=batch.association, slug='membre',
    ).first()
    if default_role and not MemberRole.all_objects.filter(
        membership=membership, role=default_role, is_active=True,
    ).exists():
        MemberRole.all_objects.create(
            association=batch.association, membership=membership,
            role=default_role, assigned_by=batch.imported_by, is_active=True,
        )

    import_row.resulting_membership = membership


def _process_invite(batch, import_row):
    """Crée une Invitation pour cette ligne (canal SMS par défaut si téléphone, sinon email)."""
    from apps.invitations.models import Invitation

    # Si déjà une invitation pending pour ce téléphone, on saute
    existing = Invitation.all_objects.filter(
        association=batch.association,
        phone=import_row.parsed_telephone,
        status=Invitation.Status.PENDING,
    ).first()
    if existing:
        import_row.resulting_invitation_id = existing.id
        return

    channel = (
        Invitation.Channel.EMAIL if import_row.parsed_email
        else Invitation.Channel.SMS
    )

    name = ' '.join(filter(None, [
        import_row.parsed_first_name, import_row.parsed_last_name,
    ])).strip()

    invitation = Invitation.all_objects.create(
        association=batch.association,
        invited_by=batch.imported_by,
        email=import_row.parsed_email or '',
        phone=import_row.parsed_telephone,
        name=name,
        channel=channel,
        message='',
        expires_at=timezone.now() + timedelta(days=7),
    )
    # Envoi best-effort (SMS/Email/WhatsApp) — gère soi-même les erreurs en aval
    try:
        from apps.invitations.services import InvitationService
        InvitationService._dispatch(invitation)  # type: ignore[attr-defined]
    except Exception:
        pass

    import_row.resulting_invitation_id = invitation.id
