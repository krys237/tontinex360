"""
Tests du parsing Excel et du workflow d'import des membres.
"""
from io import BytesIO
from django.test import TestCase
from openpyxl import Workbook

from apps.members.import_service import (
    parse_excel, validate_rows, build_template_xlsx, process_batch,
)
from apps.members.models import (
    Membership, MemberImportBatch, MemberImportRow,
)
from apps.members.tests._fixtures import TestScenario, make_membership, make_user


def _build_xlsx(rows):
    """Crée un xlsx en mémoire avec les colonnes par défaut + les rows fournies."""
    wb = Workbook()
    ws = wb.active
    ws.append(['telephone', 'first_name', 'last_name', 'email', 'member_number'])
    for r in rows:
        ws.append(r)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


class ParseExcelTests(TestCase):

    def test_parse_valid_rows(self):
        content = _build_xlsx([
            ['+237699111111', 'Jean', 'Dupont', 'jean@x.com', 'M-001'],
            ['+237699222222', 'Marie', 'NDIAYE', '', ''],
        ])
        rows, errors = parse_excel(content)
        self.assertEqual(errors, [])
        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]['parsed_telephone'], '+237699111111')
        self.assertEqual(rows[0]['parsed_first_name'], 'Jean')

    def test_parse_normalizes_phone(self):
        content = _build_xlsx([
            ['+225 07 08 09 10 11', 'A', 'B', '', ''],
        ])
        rows, _ = parse_excel(content)
        # Espaces enlevés mais + conservé
        self.assertEqual(rows[0]['parsed_telephone'], '+22507080910 11'.replace(' ', ''))

    def test_parse_accepts_french_headers(self):
        # En-têtes en français : téléphone, prénom, nom
        wb = Workbook()
        ws = wb.active
        ws.append(['Téléphone', 'Prénom', 'Nom', 'Email', 'Matricule'])
        ws.append(['+237699111111', 'Jean', 'Dupont', 'jean@x.com', 'M-001'])
        buf = BytesIO()
        wb.save(buf)

        rows, errors = parse_excel(buf.getvalue())
        self.assertEqual(errors, [])
        self.assertEqual(rows[0]['parsed_telephone'], '+237699111111')
        self.assertEqual(rows[0]['parsed_first_name'], 'Jean')

    def test_parse_skips_empty_rows(self):
        wb = Workbook()
        ws = wb.active
        ws.append(['telephone', 'first_name', 'last_name'])
        ws.append(['+237699111111', 'A', 'B'])
        ws.append([None, None, None])  # Vide → skip
        ws.append(['+237699222222', 'C', 'D'])
        buf = BytesIO()
        wb.save(buf)

        rows, errors = parse_excel(buf.getvalue())
        self.assertEqual(len(rows), 2)

    def test_parse_rejects_file_without_phone_column(self):
        wb = Workbook()
        ws = wb.active
        ws.append(['nom', 'prenom'])  # pas de telephone
        ws.append(['Dupont', 'Jean'])
        buf = BytesIO()
        wb.save(buf)
        rows, errors = parse_excel(buf.getvalue())
        self.assertEqual(rows, [])
        self.assertEqual(len(errors), 1)
        self.assertIn('telephone', errors[0].lower())


class ValidateRowsTests(TestCase):

    def setUp(self):
        self.scn = TestScenario.build_full()

    def test_validation_marks_duplicates_with_existing_members(self):
        # member1 existe déjà avec son téléphone
        existing_phone = self.scn.member1.user.telephone
        rows = [
            {'row_number': 2, 'raw_data': {}, 'parsed_telephone': existing_phone,
             'parsed_first_name': 'X', 'parsed_last_name': 'Y',
             'parsed_email': '', 'parsed_member_number': ''},
            {'row_number': 3, 'raw_data': {}, 'parsed_telephone': '+237699999999',
             'parsed_first_name': 'A', 'parsed_last_name': 'B',
             'parsed_email': '', 'parsed_member_number': ''},
        ]
        result = validate_rows(self.scn.association, rows)
        self.assertEqual(result[0]['validation'], 'duplicate')
        self.assertEqual(result[1]['validation'], 'ok')

    def test_validation_marks_file_duplicates(self):
        rows = [
            {'row_number': 2, 'raw_data': {}, 'parsed_telephone': '+237699000111',
             'parsed_first_name': 'X', 'parsed_last_name': 'Y',
             'parsed_email': '', 'parsed_member_number': ''},
            {'row_number': 3, 'raw_data': {}, 'parsed_telephone': '+237699000111',  # même
             'parsed_first_name': 'A', 'parsed_last_name': 'B',
             'parsed_email': '', 'parsed_member_number': ''},
        ]
        result = validate_rows(self.scn.association, rows)
        self.assertEqual(result[0]['validation'], 'ok')
        self.assertEqual(result[1]['validation'], 'doublon_fichier')

    def test_validation_marks_invalid_phone(self):
        rows = [
            {'row_number': 2, 'raw_data': {}, 'parsed_telephone': '',
             'parsed_first_name': 'X', 'parsed_last_name': 'Y',
             'parsed_email': '', 'parsed_member_number': ''},
        ]
        result = validate_rows(self.scn.association, rows)
        self.assertEqual(result[0]['validation'], 'invalid')


class TemplateTests(TestCase):

    def test_template_is_valid_xlsx(self):
        content = build_template_xlsx()
        from openpyxl import load_workbook
        wb = load_workbook(BytesIO(content), read_only=True)
        ws = wb.active
        # 1ère ligne = en-tête
        header = next(ws.iter_rows(values_only=True))
        self.assertIn('telephone', header)
